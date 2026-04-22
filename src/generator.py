"""
分類済みJSONからExcelファイルを生成するモジュール
"""
import json
import os
import sys
import logging
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from config import CATEGORIES, HEADER_GROUPS, OUTPUT_DIR

logger = logging.getLogger(__name__)

COLOR_HEADER_BG = "1F4E79"
COLOR_HEADER_FG = "FFFFFF"
COLOR_SUB_BG    = "BDD7EE"
COLOR_NEG_BG    = "FFE0E0"
COLOR_POS_BG    = "E2EFDA"
COLOR_TAG_VAL   = "FFC000"
COLOR_EVEN_ROW  = "F2F2F2"


def generate_excel(articles: list[dict], year: int, month: int) -> str:
    """
    分類済み記事リストからExcelを生成し、パスを返す。

    articles の各要素は以下の形式:
      {"title": str, "date": str, "source": str,
       "categories": {"negative": bool, "positive": bool, ...}}
    """
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{year}-{month:02d}"

    _write_header(ws)
    _write_data(ws, articles)
    _apply_column_widths(ws)
    ws.freeze_panes = "C3"
    ws.sheet_view.showGridLines = False

    filename = _build_filename(year, month)
    path = os.path.join(OUTPUT_DIR, filename)
    wb.save(path)
    logger.info(f"Excel保存: {path}")
    return path


def _write_header(ws):
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor=COLOR_HEADER_BG)
    header_font = Font(bold=True, color=COLOR_HEADER_FG, size=9, name="游ゴシック")
    sub_fill    = PatternFill("solid", fgColor=COLOR_SUB_BG)
    sub_font    = Font(bold=True, size=8, name="游ゴシック")
    center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for group_name, col_start, col_end in HEADER_GROUPS:
        cell = ws.cell(row=1, column=col_start, value=group_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_wrap
        cell.border = border
        if col_start != col_end:
            ws.merge_cells(
                start_row=1, start_column=col_start,
                end_row=1,   end_column=col_end,
            )

    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    ws.cell(row=1, column=2).fill   = header_fill
    ws.cell(row=1, column=2).border = border

    ws.cell(row=2, column=2, value="日付").fill      = sub_fill
    ws.cell(row=2, column=2).font      = sub_font
    ws.cell(row=2, column=2).alignment = center_wrap
    ws.cell(row=2, column=2).border    = border

    for key, (col_idx, display_name, _) in CATEGORIES.items():
        cell = ws.cell(row=2, column=col_idx, value=display_name)
        if key == "negative":
            cell.fill = PatternFill("solid", fgColor=COLOR_NEG_BG)
        elif key == "positive":
            cell.fill = PatternFill("solid", fgColor=COLOR_POS_BG)
        else:
            cell.fill = sub_fill
        cell.font      = sub_font
        cell.alignment = center_wrap
        cell.border    = border

    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 40


def _write_data(ws, articles):
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center    = Alignment(horizontal="center", vertical="center")
    left_wrap = Alignment(horizontal="left", vertical="center", wrap_text=True)
    tag_font  = Font(bold=True, color=COLOR_TAG_VAL, size=9, name="游ゴシック")
    link_font = Font(color="0563C1", underline="single", size=10, name="游ゴシック")
    plain_font = Font(size=10, name="游ゴシック")
    even_fill = PatternFill("solid", fgColor=COLOR_EVEN_ROW)

    for i, article in enumerate(articles):
        row  = i + 3
        fill = even_fill if i % 2 == 0 else PatternFill()

        cell = ws.cell(row=row, column=1, value=article["title"])
        url = article.get("url", "")
        if url:
            cell.hyperlink = url
            cell.font = link_font
        else:
            cell.font = plain_font
        cell.alignment = left_wrap
        cell.border    = border
        cell.fill      = fill

        cell = ws.cell(row=row, column=2, value=article["date"])
        cell.alignment = center
        cell.border    = border
        cell.fill      = fill

        for key, (col_idx, _, _) in CATEGORIES.items():
            cell = ws.cell(row=row, column=col_idx)
            if article.get("categories", {}).get(key):
                cell.value     = 1
                cell.font      = tag_font
                cell.alignment = center
            cell.border = border
            cell.fill   = fill


def _apply_column_widths(ws):
    ws.column_dimensions["A"].width = 80
    ws.column_dimensions["B"].width = 11
    for col_idx in range(3, 27):
        ws.column_dimensions[get_column_letter(col_idx)].width = 7


def _build_filename(year, month):
    return f"MarketNews-{year}{month:02d}.xlsx"


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")

    if len(sys.argv) == 3:
        year, month = int(sys.argv[1]), int(sys.argv[2])
    else:
        now = datetime.now()
        year, month = now.year, now.month

    json_path = f"data/classified_{year}{month:02d}.json"
    if not os.path.exists(json_path):
        print(f"エラー: {json_path} が見つかりません")
        sys.exit(1)

    with open(json_path, encoding="utf-8") as f:
        articles = json.load(f)

    path = generate_excel(articles, year, month)
    print(f"Excel生成完了: {path} ({len(articles)}件)")
